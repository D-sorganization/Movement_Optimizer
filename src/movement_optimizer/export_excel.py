# Copyright (c) 2026 D-Sorganization. All rights reserved.
"""Excel export for optimization results.

Design Principles:
    DBC  -- preconditions checked at function entry.
    DRY  -- sheet-writing helpers avoid duplicated loop logic.

Requires the optional ``openpyxl`` package.  The function raises
``ImportError`` with a helpful message when it is not installed.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import TYPE_CHECKING

import numpy as np

if TYPE_CHECKING:
    from .trajectory.result import OptimizationResult

logger = logging.getLogger(__name__)


def _validate_path(path: str | Path) -> Path:
    """Normalize and basic-validate the output path.

    Raises:
        ValueError: If ``path`` contains a null byte.
    """
    p = Path(path)
    if "\x00" in str(p):
        raise ValueError("path must not contain null bytes")
    return p


def _write_summary_sheet(
    ws,
    result: OptimizationResult,
    exercise_name: str,
    body_mass_kg: float | None,
    body_height_m: float | None,
) -> None:
    """Populate the Summary worksheet."""
    rows: list[tuple[str, object]] = [
        ("Exercise", exercise_name or "-"),
        ("Converged", "Yes" if result.success else "No"),
        ("Cost", round(float(result.cost), 6)),
        ("Duration (s)", round(float(result.t[-1] - result.t[0]), 4)),
        ("Samples (N)", len(result.t)),
        ("COM horizontal range (cm)", round(float(result.com_horizontal_range_cm), 4)),
        ("Elapsed time (s)", round(float(result.elapsed_s), 3)),
        ("Cost evaluations", int(result.n_evals)),
        ("Joint limit violations", int(result.n_joint_limit_violations)),
    ]
    if body_mass_kg is not None:
        rows.insert(1, ("Body mass (kg)", round(float(body_mass_kg), 2)))
    if body_height_m is not None:
        rows.insert(2, ("Body height (m)", round(float(body_height_m), 3)))

    for label, value in rows:
        ws.append([label, value])

    ws.append([])  # blank separator

    joint_labels = ["Ankle (joint 1)", "Knee (joint 2)", "Hip (joint 3)"]
    ws.append(["Joint torque statistics", "Peak |tau| (N*m)", "Mean |tau| (N*m)", "RMS tau (N*m)"])
    n_dof = result.torques.shape[1]
    for j in range(n_dof):
        col = result.torques[:, j]
        peak = float(np.max(np.abs(col)))
        mean = float(np.mean(np.abs(col)))
        rms = float(np.sqrt(np.mean(col**2)))
        label = joint_labels[j] if j < len(joint_labels) else f"Joint {j + 1}"
        ws.append([label, round(peak, 3), round(mean, 3), round(rms, 3)])


def _write_trajectory_sheet(ws, result: OptimizationResult) -> None:
    """Populate the Trajectory worksheet with time-series joint kinematics."""
    n = len(result.t)
    n_dof = result.q.shape[1]
    header = (
        ["time (s)"]
        + [f"q{j + 1} (deg)" for j in range(n_dof)]
        + [f"dq{j + 1} (deg/s)" for j in range(n_dof)]
    )
    ws.append(header)
    for i in range(n):
        row: list[object] = [round(float(result.t[i]), 6)]
        row += [round(float(np.degrees(result.q[i, j])), 4) for j in range(n_dof)]
        row += [round(float(np.degrees(result.qd[i, j])), 4) for j in range(n_dof)]
        ws.append(row)


def _write_torques_sheet(ws, result: OptimizationResult) -> None:
    """Populate the Torques worksheet with time-series joint torques."""
    n = len(result.t)
    n_dof = result.torques.shape[1]
    header = ["time (s)"] + [f"tau{j + 1} (N*m)" for j in range(n_dof)]
    ws.append(header)
    for i in range(n):
        row: list[object] = [round(float(result.t[i]), 6)]
        row += [round(float(result.torques[i, j]), 4) for j in range(n_dof)]
        ws.append(row)


def export_to_excel(
    result: OptimizationResult,
    path: str | Path,
    exercise_name: str = "",
    body_mass_kg: float | None = None,
    body_height_m: float | None = None,
) -> None:
    """Export optimization result to an Excel workbook with multiple sheets.

    Creates three sheets:

    * **Summary** - key parameters and per-joint statistics (peak torque,
      mean torque, RMS torque) plus convergence status.
    * **Trajectory** - time-series columns ``time``, ``q1``-``q3`` (degrees),
      ``dq1``-``dq3`` (deg/s).
    * **Torques** - time-series columns ``time``, ``tau1``-``tau3`` (N*m).

    Args:
        result: Completed :class:`~movement_optimizer.trajectory.OptimizationResult`.
        path: Output file path (should end in ``.xlsx``).
        exercise_name: Human-readable exercise label written to the Summary sheet.
        body_mass_kg: Optional body mass written to Summary sheet.
        body_height_m: Optional body height written to Summary sheet.

    Raises:
        ImportError: If ``openpyxl`` is not installed.
        ValueError: If ``result`` is ``None``, ``path`` contains a null byte, or
            the trajectory arrays have incompatible shapes.
    """
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel export. Install it with: pip install openpyxl"
        ) from exc

    if result is None:
        raise ValueError("result must not be None")

    safe_path = _validate_path(path)

    n = len(result.t)
    if result.q.shape[0] != n or result.torques.shape[0] != n:
        raise ValueError(
            f"Trajectory array length mismatch: t has {n} rows but "
            f"q has {result.q.shape[0]} and torques has {result.torques.shape[0]}"
        )

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(ws_summary, result, exercise_name, body_mass_kg, body_height_m)

    ws_traj = wb.create_sheet("Trajectory")
    _write_trajectory_sheet(ws_traj, result)

    ws_torques = wb.create_sheet("Torques")
    _write_torques_sheet(ws_torques, result)

    safe_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(safe_path))
    logger.info(
        "Exported Excel workbook to %s (%d samples, %d DOF)",
        safe_path,
        n,
        result.torques.shape[1],
    )
