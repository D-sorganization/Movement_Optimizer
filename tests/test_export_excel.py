# Copyright (c) 2026 D-Sorganization. All rights reserved.
"""Tests for export_excel module."""

from __future__ import annotations

import pytest

pytest.importorskip("openpyxl")

import openpyxl
from conftest import make_test_result

from movement_optimizer.export_excel import export_to_excel


class TestExportToExcel:
    def test_creates_file_with_expected_sheets(self, tmp_path):
        """export_to_excel creates workbook sheets for summary, data, and statistics."""
        result = make_test_result()
        path = tmp_path / "out.xlsx"

        export_to_excel(result, path)

        assert path.exists()
        wb = openpyxl.load_workbook(str(path))
        assert set(wb.sheetnames) == {"Summary", "Trajectory", "Torques", "Statistics"}

    @pytest.mark.timeout(180)
    def test_summary_sheet_has_non_empty_data(self, tmp_path):
        """The Summary sheet contains at least one row of non-empty data."""
        result = make_test_result()
        path = tmp_path / "summary_check.xlsx"

        export_to_excel(result, path, exercise_name="Squat", body_mass_kg=75.0)

        wb = openpyxl.load_workbook(str(path))
        ws = wb["Summary"]
        non_empty_rows = [
            row for row in ws.iter_rows(values_only=True) if any(v is not None for v in row)
        ]
        assert len(non_empty_rows) > 0

    def test_trajectory_sheet_column_count(self, tmp_path):
        """Trajectory sheet has time + q + dq columns (1 + 3 + 3 = 7 for 3-DOF)."""
        result = make_test_result()
        path = tmp_path / "traj_cols.xlsx"

        export_to_excel(result, path)

        wb = openpyxl.load_workbook(str(path))
        ws = wb["Trajectory"]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        # 1 time + 3 q + 3 dq = 7
        assert len([h for h in header if h is not None]) == 7

    def test_torques_sheet_column_count(self, tmp_path):
        """Torques sheet has time + tau columns (1 + 3 = 4 for 3-DOF)."""
        result = make_test_result()
        path = tmp_path / "torques_cols.xlsx"

        export_to_excel(result, path)

        wb = openpyxl.load_workbook(str(path))
        ws = wb["Torques"]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        # 1 time + 3 tau = 4
        assert len([h for h in header if h is not None]) == 4

    def test_none_result_raises_value_error(self, tmp_path):
        """ValueError is raised when result is None."""
        with pytest.raises(ValueError, match="result must not be None"):
            export_to_excel(None, str(tmp_path / "out.xlsx"))

    def test_null_byte_in_path_raises_value_error(self, tmp_path):
        """ValueError is raised when path contains a null byte."""
        result = make_test_result()
        with pytest.raises(ValueError, match="null bytes"):
            export_to_excel(result, "out\x00.xlsx")

    def test_creates_parent_directories(self, tmp_path):
        """export_to_excel creates missing parent directories automatically."""
        result = make_test_result()
        path = tmp_path / "subdir" / "nested" / "out.xlsx"

        export_to_excel(result, path)

        assert path.exists()

    def test_optional_metadata_written_to_summary(self, tmp_path):
        """Body mass and height appear in the Summary sheet when provided."""
        result = make_test_result()
        path = tmp_path / "meta.xlsx"

        export_to_excel(
            result, path, exercise_name="Deadlift", body_mass_kg=80.0, body_height_m=1.82
        )

        wb = openpyxl.load_workbook(str(path))
        ws = wb["Summary"]
        all_values = [cell for row in ws.iter_rows(values_only=True) for cell in row]
        assert "Deadlift" in all_values
        assert 80.0 in all_values
        assert 1.82 in all_values

    def test_statistics_sheet_contains_required_metrics(self, tmp_path):
        """The Statistics sheet includes mean, std, min, max, and recommendations."""
        result = make_test_result()
        path = tmp_path / "stats.xlsx"

        export_to_excel(result, path)

        wb = openpyxl.load_workbook(str(path))
        ws = wb["Statistics"]
        values = [cell for row in ws.iter_rows(values_only=True) for cell in row if cell]
        assert "Mean (N*m)" in values
        assert "Std dev (N*m)" in values
        assert "Min (N*m)" in values
        assert "Max (N*m)" in values
        assert "Recommendations" in values
